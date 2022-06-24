# -*- coding: UTF-8 -*-
import os
import time
import urllib
from datetime import datetime
import threading
from lxml import etree
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup

header = {
    "cookie": "bbs_token=P62jCVsWr9504BF_2F7Iq0cjSsOikTV423qIdkfuTJlXEg_2Fp0eE4MD3MHQ6nw_2Fzk_2FBTyw8_2FxFV5iulMqNjAcyOzA_3D_3D; bbs_sid=458013c9035834363c49487a843e3def; UM_distinctid=17c7c5991345db-070396d0205a9e-513c1e45-32a000-17c7c59913583f; atarget=n; CNZZDATA1277689586=267366360-1634168641-|1637406544",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36 Edg/93.0.961.38",
}
session = requests.session()
requests.packages.urllib3.disable_warnings()

ret = []

baseurl = "https://www.scboy.cc/?user-{}.htm"

wb = Workbook()

ws = wb.active

# 表头
ws["A1"] = ""
ws["B1"] = "ID"
ws["C1"] = "用户名"
ws["D1"] = "主题数"
ws["E1"] = "贴子数"
ws["F1"] = "粉丝数"
ws["G1"] = "积分"
ws["H1"] = "金币"
ws["I1"] = "精华"
ws["J1"] = "用户组"
ws["K1"] = "创建时间"

try:

    for x in range(9484, 50000):

        time.sleep(0.1)

        url = str(baseurl).format(x)

        res = session.get(url, headers=header, verify=False)

        res.encoding = "utf-8"

        reslxml = BeautifulSoup(res.text, "lxml")
        # 用户ID
        userid = x
        # 用户名
        usernameselecter = "#user_main > div.card > div.card-body > div > div.col-md-2.col-sm-12.text-center > b"

        leftselecter = (
            "#user_main > div.card > div.card-body > div > div.col-4.col-sm-5"
        )
        rightselecter = "#user_main > div.card > div.card-body > div > div.col-auto"
        username = ""
        try:
            username = reslxml.select(usernameselecter)[0].text
            username = username.replace("\n", "")
            username = username.replace("\t", "")
            username = username.replace("\r", "")
            username = username.replace(" ", "")
        except:
            print(userid + "名称获取失败")

        leftresults = ""
        try:
            leftresults = str(reslxml.select(leftselecter)[0].text)
            leftresults = leftresults.replace("\n", "")
            leftresults = leftresults.replace("\t", "")
            leftresults = leftresults.replace("\r", "")
            leftresults = leftresults.replace(" ", "")

        except:
            print()
        rightresults = ""
        try:
            rightresults = str(reslxml.select(rightselecter)[0].text)
            rightresults = rightresults.replace("\n", "")
            rightresults = rightresults.replace("\t", "")
            rightresults = rightresults.replace("\r", "")
            rightresults = rightresults.replace(" ", "")
        except:
            print()

        try:
            print()
        except:
            print()

        lefts = leftresults.split("：")
        rights = rightresults.split("：")
        # print(lefts)
        # print(rights)

        # ID 用户名 主题数 贴子数 粉丝数 积分 金币 精华 用户组 创建时间

        item = []
        item.append("")
        try:
            item.append(str(x))
            item.append(username)
            item.append((lefts[1])[0:-3])
            item.append((lefts[2])[0:-3])
            item.append((lefts[3])[0:-2])
            item.append((lefts[4])[0:-2])
            item.append((lefts[5])[0:-3])
            item.append((lefts[6]))
            item.append((rights[1])[0:-4])
            item.append((rights[2])[0:-4])

            ws.append(item)
        except:
            ws.append(item)
        print(item)
except:
    # 保存
    wb.save("scboy.xlsx")
wb.save("scboy.xlsx")
